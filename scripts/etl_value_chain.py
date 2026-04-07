"""
ETL: AI Infrastructure Value Chain → SQLite + CSV
Extracts the Full_Universe sheet (stock-level data with supply chain taxonomy)
and the AI Value Chain sheet (taxonomy structure with returns).
"""

import sqlite3
import pandas as pd
from pathlib import Path

RAW_DIR = Path(__file__).parent.parent / "data" / "raw"
DB_PATH = Path(__file__).parent.parent / "data" / "db" / "ai_research.db"
CSV_DIR = Path(__file__).parent.parent / "data" / "processed"

VC_FILE = RAW_DIR / "AI Infrastructure Value Chain - Expanded.xlsx"

# Sub-sector sheets and their segment/sub-bucket classification
SUBSECTOR_SHEETS = [
    ("Semi Production - IC Design", "Semi Production", "IC Design"),
    ("Semi Production - OSAT", "Semi Production", "OSAT"),
    ("Semi Production - Foundry", "Semi Production", "Foundry"),
    ("Semi Production - IDM (semi w-", "Semi Production", "IDM"),
    ("Semi Production - Semi Design S", "Semi Production", "Semi Design Services"),
    ("Semi Production - Semi Capital ", "Semi Production", "Semi Capital Equipment"),
    ("Processor - GPU", "Processor", "GPU"),
    ("Processor - CPU", "Processor", "CPU"),
    ("Server Components - Power Suppl", "Server Components", "Power Supply"),
    ("Server Components - Passive Com", "Server Components", "Passive Component"),
    ("Server Components - Thermal Sol", "Server Components", "Thermal Solution"),
    ("Server Components - PCB", "Server Components", "PCB"),
    ("Server - Server Brands", "Server", "Server Brands"),
    ("Server - ODM - EMS", "Server", "ODM / EMS"),
    ("Network - InfiniBand", "Network", "InfiniBand"),
    ("Network - Ethernet", "Network", "Ethernet"),
    ("Network - DCI (Routing-Optical)", "Network", "DCI (Routing/Optical)"),
    ("Network - Memory-Storage", "Network", "Memory/Storage"),
    ("Network - Cabling", "Network", "Cabling"),
    ("Internal Power-Cooling - Liquid", "Internal Power/Cooling", "Liquid Cooling"),
    ("Internal Power-Cooling - Power ", "Internal Power/Cooling", "Power Electronics"),
    ("Internal Power-Cooling - Uninte", "Internal Power/Cooling", "UPS"),
    ("Power Supply - Grid + Onsite Re", "Power Supply", "Grid + Onsite Renewable Storage"),
    ("Power Supply - Generators", "Power Supply", "Generators & Grid Connection"),
    ("Power Supply - Grid Infrastruct", "Power Supply", "Grid Infrastructure"),
    ("Power Supply - Behind Fence @ N", "Power Supply", "Nuclear"),
    ("Power Supply - Fuel Cells", "Power Supply", "Fuel Cells"),
    ("Owners-Operators - Hyperscalers", "Owners/Operators", "Hyperscalers"),
    ("Owners-Operators - DC REITs", "Owners/Operators", "DC REITs"),
    ("Owners-Operators - PE", "Owners/Operators", "Private Equity"),
    ("Owners-Operators - Enterprises-", "Owners/Operators", "Enterprises/Tier 2"),
    ("Owners-Operators - Neo-cloud", "Owners/Operators", "Neo-cloud"),
]


def parse_full_universe(wb):
    """Parse the Full_Universe sheet into a clean DataFrame."""
    ws = wb["Full_Universe"]

    # Find header row (contains "Co. name")
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=15):
        for cell in row:
            if cell.value and str(cell.value).strip() == "Co. name":
                header_row = cell.row
                break
        if header_row:
            break

    if not header_row:
        raise ValueError("Could not find header row in Full_Universe")

    # Read headers
    headers = []
    for cell in list(ws.iter_rows(min_row=header_row, max_row=header_row))[0]:
        headers.append(str(cell.value).strip() if cell.value else f"col_{len(headers)}")

    # Read data rows
    records = []
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        vals = [cell.value for cell in row]
        if vals[0] and str(vals[0]).strip():  # has company name
            records.append(dict(zip(headers, vals)))

    df = pd.DataFrame(records)

    # Clean column names
    col_map = {
        "Co. name": "company",
        "Ticker": "ticker",
        "Analyst region": "region",
        "Industry": "industry",
        "Segment": "segment",
        "Sub-bucket": "sub_bucket",
        "Included": "included",
        "Upside to PT": "upside_to_pt",
        "Downside to Bear": "downside_to_bear",
        "Upside to Bull": "upside_to_bull",
        "Bull:Bear Skew": "bull_bear_skew",
        "3M Probability, Base Case (>)": "prob_3m_base",
        "3M Probability, Bull Case (>)": "prob_3m_bull",
        "12M Probability, Base Case (>)": "prob_12m_base",
        "12M Probability, Bull Case (>)": "prob_12m_bull",
    }

    rename_cols = {k: v for k, v in col_map.items() if k in df.columns}
    df = df.rename(columns=rename_cols)
    keep = [v for v in col_map.values() if v in df.columns]
    df = df[keep]

    # Convert numeric columns
    numeric_cols = [
        "upside_to_pt", "downside_to_bear", "upside_to_bull",
        "bull_bear_skew", "prob_3m_base", "prob_3m_bull",
        "prob_12m_base", "prob_12m_bull",
    ]
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # Clean included flag
    if "included" in df.columns:
        df["included"] = df["included"].apply(
            lambda x: True if x is True or str(x).strip().lower() == "true" else False
        )

    return df


def build_taxonomy():
    """Build the supply chain taxonomy from sheet names."""
    records = []
    for _, segment, sub_bucket in SUBSECTOR_SHEETS:
        records.append({
            "level": "Inside DC" if segment in (
                "Semi Production", "Processor", "Server Components",
                "Server", "Network", "Internal Power/Cooling"
            ) else "Outside DC",
            "segment": segment,
            "sub_bucket": sub_bucket,
        })
    return pd.DataFrame(records)


def run():
    print(f"Reading {VC_FILE}...")
    wb = pd.ExcelFile(VC_FILE)

    # Parse Full Universe
    import openpyxl
    wb_ox = openpyxl.load_workbook(VC_FILE, read_only=True, data_only=True)
    df_universe = parse_full_universe(wb_ox)
    wb_ox.close()

    print(f"  Full Universe: {len(df_universe)} stocks ({df_universe['included'].sum()} included)")

    # Build taxonomy
    df_taxonomy = build_taxonomy()
    print(f"  Taxonomy: {len(df_taxonomy)} sub-sectors across {df_taxonomy['segment'].nunique()} segments")

    # Write to SQLite
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    df_universe.to_sql("value_chain_universe", conn, if_exists="replace", index=False)
    df_taxonomy.to_sql("value_chain_taxonomy", conn, if_exists="replace", index=False)

    # Create joined view with mapping data
    conn.execute("DROP VIEW IF EXISTS v_full_universe")
    conn.execute("""
        CREATE VIEW v_full_universe AS
        SELECT
            vc.company, vc.ticker, vc.region, vc.industry,
            vc.segment, vc.sub_bucket, vc.included,
            vc.upside_to_pt, vc.downside_to_bear, vc.upside_to_bull,
            vc.bull_bear_skew,
            m.exposure_latest, m.materiality_latest, m.pricing_power_latest,
            m.materiality_score, m.pricing_power_score, m.materiality_trend,
            m.alpha_flag, m.gics_sector
        FROM value_chain_universe vc
        LEFT JOIN mapping m ON vc.ticker = m.ticker
    """)

    conn.commit()
    conn.close()

    # Write CSVs
    CSV_DIR.mkdir(parents=True, exist_ok=True)
    df_universe.to_csv(CSV_DIR / "value_chain_universe.csv", index=False)
    df_taxonomy.to_csv(CSV_DIR / "value_chain_taxonomy.csv", index=False)
    print(f"  Written to {DB_PATH} and {CSV_DIR}")


if __name__ == "__main__":
    run()
